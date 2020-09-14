from openpyxl import Workbook
from openpyxl.styles import Font
import Tools.gl as gl
from common import *

IN_DIR = 'C:/Py/OUT/CAPC5_OUT_20200911/'
OUT_DIR = 'C:/Py/OUT/test.xlsx'
CSV_DIR = 'C:/Py/OUT/CAPC5_20200911.csv'
TXT_CONVERT = ['PRM', 'PDS', 'POINT', 'IDPDL', 'RAE', 'CCD']

def csv_to_xls_folder(in_dir = IN_DIR):
	
	file_list = get_file_list(IN_DIR)
	n = len(file_list)
	log("Conversion des {} fichiers du dossier {} au format xlsx".format(n, IN_DIR))
	print_com('|')
	i = 0
	for elt in file_list:
		i += 1
		in_dir = IN_DIR + elt
		out_dir = in_dir.replace('.csv', '.xlsx')
		log("Conversion du fichier {}...".format(elt))
		csv_to_xls(in_dir, out_dir)
		log("Conversion du fichier {} terminée, {} fichiers convertis, {} fichiers restants".format(elt, i, n-i))
	
	dur = get_duration_ms(gl.start_time)	
	finish(dur)

def csv_to_xls(in_dir = IN_DIR, out_dir = OUT_DIR):
	
	# Création du classeur
	wb = Workbook()
	ws = wb.active
	ws.title = "DATA"
	
	# Chargement du fichier csv
	csv = load_csv(in_dir)
	header = csv[0]
	
	# Remplissage du fichier excel
	i = 0
	for row in csv:
		j = 0
		i += 1
		for elt in row:
			j += 1
			c = ws.cell(row=i, column=j, value=elt)
			if header[j-1] in TXT_CONVERT:
				c.number_format = '@'
	
	# Première ligne en gras
	j = 0
	for elt in header:
		j += 1
		c = ws.cell(row=1, column=j)
		c.font = Font(bold=True)
	
	# Première ligne gelée
	c = ws['A2']
	ws.freeze_panes = c
	
	# Sauvegarde du fichier
	wb.save(out_dir)
	
def finish(dur):

	print_com('|')
	s = "Conversion terminé en {}."
	s = s.format(get_duration_string(dur))
	log(s)
	send_notif(s, "Conversion XLS", dur)